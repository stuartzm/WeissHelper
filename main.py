from openpyxl import Workbook
from openpyxl import load_workbook

########################################
########################################
## MAKE SURE THE FILE IS CLOSED FIRST ##
## MAKE SURE THE FILE IS CLOSED FIRST ##
## MAKE SURE THE FILE IS CLOSED FIRST ##
########################################
########################################

#Enter the file name and the extension here:
FILENAME = input("Enter the file name and the extension here: ")

wb = load_workbook(FILENAME)
ws = wb.active

FIRSTROW_COLUMN_LETTER = input("The column with the first set of info: ")
LASTROW_COLUMN_LETTER = input("The column with the last set of info: ")

COLUMN_TO_WRITE_TO = input("The column that you would like to write info to: ")
COLUMN_END_WRITE_TO = input("The ending column that you would like to write info to: ")
ROW_TO_START_ON = input("This number should be the first row number of the info: ")
ROW_TO_END_ON = input("This number should be the last row number of the info: ")

ItemsUpDown = input("The amount of items in a column: ")
ItemsLeftRight = input("The amount of items in a row: ")

INTERVAL = input("The interval between each set of rows: ")

write_row = int(ROW_TO_START_ON)
for curr_interval in range(0, int(ItemsUpDown)):
	if curr_interval > int(ItemsUpDown):
		continue

	current = int(ROW_TO_START_ON) + curr_interval
	adder = current

	while adder <= int(ROW_TO_END_ON):
		for item, placer in zip(ws[FIRSTROW_COLUMN_LETTER+str(adder):LASTROW_COLUMN_LETTER+str(adder)], ws[COLUMN_TO_WRITE_TO+str(write_row):COLUMN_END_WRITE_TO+str(write_row)]):
			for cell, cell2 in zip(item, placer):
				cell2.value = cell.value
			write_row += 1
		adder = adder + int(INTERVAL)
	write_row += 3

wb.save(FILENAME)

